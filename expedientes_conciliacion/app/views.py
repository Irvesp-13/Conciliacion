from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Q
from .models import *
from .models import Empleado
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from .models import CargaDescarga
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

def iniciar_sesion(request):
    if request.method == 'POST':
        nombre = request.POST['nombre']
        clave_empleado = request.POST['clave_empleado']
        
        try:
            # Buscar al empleado por su clave de empleado y nombre
            empleado = Empleado.objects.get(clave_empleado=clave_empleado, nombre=nombre)
            
            # Guardar el ID del empleado en la sesión
            request.session['empleado_id'] = empleado.id
            return redirect('bienvenida')
        except Empleado.DoesNotExist:
            # Mostrar mensaje de error si las credenciales son incorrectas
            return render(request, 'iniciar_sesion.html', {'error': 'Credenciales incorrectas'})
    
    return render(request, 'iniciar_sesion.html')

def bienvenida(request):
    # Verificar si el empleado está autenticado
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        return redirect('iniciar_sesion')
    
    # Obtener el empleado autenticado
    empleado = Empleado.objects.get(id=empleado_id)
    
    # Obtener registros de la tabla dosmilsiete
    expedientes = DosMilSiete.objects.all()
    
    return render(request, 'bienvenida.html', {
        'empleado': empleado,
        'expedientes': expedientes,
    })


def cerrar_sesion(request):
    # Eliminar el ID del empleado de la sesión
    if 'empleado_id' in request.session:
        del request.session['empleado_id']
    return redirect('iniciar_sesion')

def agregar_expediente(request):
    # Verificar si el usuario es administrador
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        return redirect('iniciar_sesion')
    
    empleado = Empleado.objects.get(id=empleado_id)
    if not empleado.es_administrador():
        messages.error(request, 'No tienes permiso para realizar esta acción.')
        return redirect('bienvenida')
    
    if request.method == 'POST':
        # Get the last id_expediente and increment it
        last_expediente = DosMilSiete.objects.order_by('-id_expediente').first()
        new_id = 1 if not last_expediente else last_expediente.id_expediente + 1
        
        nuevo_expediente = DosMilSiete(
            id_expediente=new_id,  # Add this line
            expediente=request.POST['expediente'],
            junta=request.POST['junta'],
            actor=request.POST['actor'],
            demandado=request.POST['demandado'],
            no_se_ha_notificao_a_las_partes=request.POST['no_se_ha_notificao_a_las_partes'],
            empl_no_realizado=request.POST['empl_no_realizado'],
            empl_exhorto=request.POST['empl_exhorto'],
            cde=request.POST['cde'],
            oap=request.POST['oap'],
            desahogo_pruebas=request.POST['desahogo_pruebas'],
            cierre=request.POST['cierre'],
            laudo_dictado=request.POST['laudo_dictado'],
            auto_ejecucion=request.POST['auto_ejecucion'],
            prescripcion=request.POST['prescripcion'],
            regularizar=request.POST['regularizar']
        )
        nuevo_expediente.save()
        messages.success(request, 'Expediente agregado correctamente.')
        return redirect('bienvenida')
    
    return render(request, 'agregar_expediente.html')


def editar_expediente(request, id_expediente):
    if not request.session.get('empleado_id'):
        return redirect('iniciar_sesion')
    
    empleado = Empleado.objects.get(id=request.session['empleado_id'])
    if not empleado.es_administrador():
        messages.error(request, 'No tienes permiso para realizar esta acción.')
        return redirect('bienvenida')

    try:
        expediente = get_object_or_404(DosMilSiete, id_expediente=id_expediente)
        
        if request.method == 'POST':
            # Update fields matching DosMilSiete model
            expediente.expediente = request.POST.get('expediente')
            expediente.junta = request.POST.get('junta')
            expediente.actor = request.POST.get('actor')
            expediente.demandado = request.POST.get('demandado')
            expediente.no_se_ha_notificao_a_las_partes = request.POST.get('no_se_ha_notificao_a_las_partes')
            expediente.empl_no_realizado = request.POST.get('empl_no_realizado')
            expediente.empl_exhorto = request.POST.get('empl_exhorto')
            expediente.cde = request.POST.get('cde')
            expediente.oap = request.POST.get('oap')
            expediente.desahogo_pruebas = request.POST.get('desahogo_pruebas')
            expediente.cierre = request.POST.get('cierre')
            expediente.laudo_dictado = request.POST.get('laudo_dictado')
            expediente.auto_ejecucion = request.POST.get('auto_ejecucion')
            expediente.prescripcion = request.POST.get('prescripcion')
            expediente.regularizar = request.POST.get('regularizar')
            
            expediente.save()
            messages.success(request, 'Expediente actualizado correctamente.')
            return redirect('bienvenida')
        
        return render(request, 'editar_expediente.html', {'expediente': expediente})
        
    except DosMilSiete.DoesNotExist:
        messages.error(request, 'El expediente no existe.')
        return redirect('bienvenida')


def crear_empleado(request):
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        return redirect('iniciar_sesion')
    empleado = Empleado.objects.get(id=empleado_id)
    if not empleado.es_administrador():
        messages.error(request, 'No tienes permiso para acceder a esta página.')
        return redirect('bienvenida')

    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        clave_empleado = request.POST.get('clave_empleado')
        puesto = request.POST.get('puesto')
        if nombre and clave_empleado and puesto:
            if Empleado.objects.filter(clave_empleado=clave_empleado).exists():
                messages.error(request, 'La clave de empleado ya existe.')
            else:
                Empleado.objects.create(
                    nombre=nombre,
                    clave_empleado=clave_empleado,
                    puesto=puesto
                )
                messages.success(request, 'Empleado creado exitosamente.')
                return redirect('crear_empleado')
        else:
            messages.error(request, 'Todos los campos son obligatorios.')

    empleados = Empleado.objects.all()
    return render(request, 'crear_empleado.html', {'empleados': empleados})


def editar_expediente(request, id_expediente):
    if not request.session.get('empleado_id'):
        return redirect('iniciar_sesion')
    
    empleado = Empleado.objects.get(id=request.session['empleado_id'])
    if not empleado.es_administrador():
        messages.error(request, 'No tienes permiso para realizar esta acción.')
        return redirect('bienvenida')

    try:
        expediente = get_object_or_404(DosMilSiete, id_expediente=id_expediente)
        
        if request.method == 'POST':
            # Update fields matching DosMilSiete model
            expediente.expediente = request.POST.get('expediente')
            expediente.junta = request.POST.get('junta')
            expediente.actor = request.POST.get('actor')
            expediente.demandado = request.POST.get('demandado')
            expediente.no_se_ha_notificao_a_las_partes = request.POST.get('no_se_ha_notificao_a_las_partes')
            expediente.empl_no_realizado = request.POST.get('empl_no_realizado')
            expediente.empl_exhorto = request.POST.get('empl_exhorto')
            expediente.cde = request.POST.get('cde')
            expediente.oap = request.POST.get('oap')
            expediente.desahogo_pruebas = request.POST.get('desahogo_pruebas')
            expediente.cierre = request.POST.get('cierre')
            expediente.laudo_dictado = request.POST.get('laudo_dictado')
            expediente.auto_ejecucion = request.POST.get('auto_ejecucion')
            expediente.prescripcion = request.POST.get('prescripcion')
            expediente.regularizar = request.POST.get('regularizar')
            
            expediente.save()
            messages.success(request, 'Expediente actualizado correctamente.')
            return redirect('bienvenida')
        
        return render(request, 'editar_expediente.html', {'expediente': expediente})
        
    except DosMilSiete.DoesNotExist:
        messages.error(request, 'El expediente no existe.')
        return redirect('bienvenida')


def crear_empleado(request):
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        return redirect('iniciar_sesion')
    empleado = Empleado.objects.get(id=empleado_id)
    if not empleado.es_administrador():
        messages.error(request, 'No tienes permiso para acceder a esta página.')
        return redirect('bienvenida')

    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        clave_empleado = request.POST.get('clave_empleado')
        puesto = request.POST.get('puesto')
        if nombre and clave_empleado and puesto:
            if Empleado.objects.filter(clave_empleado=clave_empleado).exists():
                messages.error(request, 'La clave de empleado ya existe.')
            else:
                Empleado.objects.create(
                    nombre=nombre,
                    clave_empleado=clave_empleado,
                    puesto=puesto
                )
                messages.success(request, 'Empleado creado exitosamente.')
                return redirect('crear_empleado')
        else:
            messages.error(request, 'Todos los campos son obligatorios.')

    empleados = Empleado.objects.all()
    return render(request, 'crear_empleado.html', {'empleados': empleados})


def eliminar_expediente(request, id_expediente):
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        return redirect('iniciar_sesion')
    
    empleado = Empleado.objects.get(id=empleado_id)
    if not empleado.es_administrador():
        messages.error(request, 'No tienes permiso para eliminar expedientes.')
        return redirect('bienvenida')
    
    try:
        expediente = DosMilSiete.objects.get(id_expediente=id_expediente)
        expediente.delete()
        messages.success(request, 'Expediente eliminado correctamente.')
    except DosMilSiete.DoesNotExist:
        messages.error(request, 'El expediente no existe.')
    
    return redirect('bienvenida')


def editar_empleado(request, id):
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        return redirect('iniciar_sesion')
    admin = Empleado.objects.get(id=empleado_id)
    if not admin.es_administrador():
        messages.error(request, 'No tienes permiso para acceder a esta página.')
        return redirect('bienvenida')

    empleado = Empleado.objects.get(id=id)
    if request.method == 'POST':
        empleado.nombre = request.POST.get('nombre')
        empleado.clave_empleado = request.POST.get('clave_empleado')
        empleado.puesto = request.POST.get('puesto')
        empleado.save()
        messages.success(request, 'Empleado actualizado correctamente.')
        return redirect('crear_empleado')
    return render(request, 'editar_empleado.html', {'empleado': empleado})

def eliminar_empleado(request, id):
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        return redirect('iniciar_sesion')
    admin = Empleado.objects.get(id=empleado_id)
    if not admin.es_administrador():
        messages.error(request, 'No tienes permiso para acceder a esta página.')
        return redirect('bienvenida')

    # Prevent self-deletion
    if int(id) == int(empleado_id):
        messages.error(request, 'No puedes eliminar tu propio usuario mientras estás conectado.')
        return redirect('crear_empleado')

    empleado = Empleado.objects.get(id=id)
    empleado.delete()
    messages.success(request, 'Empleado eliminado correctamente.')
    return redirect('crear_empleado')


def cargar_expediente(request):
    if request.method == 'POST':
        empleado_id = request.session.get('empleado_id')
        expediente_id = request.POST.get('expediente_id')
        nombre_carga = request.POST.get('nombre_carga')

        try:
            empleado = Empleado.objects.get(id=empleado_id)
            expediente = DosMilSiete.objects.get(id_expediente=expediente_id)
            
            carga = CargaDescarga(
                empleado=empleado,
                expediente=expediente,
                nombre_carga=nombre_carga
            )
            carga.save()
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

def ver_cargas(request):
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        return redirect('iniciar_sesion')
    
    empleado = Empleado.objects.get(id=empleado_id)
    if not empleado.es_administrador():
        messages.error(request, 'No tienes permiso para ver esta página.')
        return redirect('bienvenida')
    
    cargas = CargaDescarga.objects.all().order_by('-fecha')
    return render(request, 'ver_cargas.html', {'cargas': cargas})


# Remove @login_required decorator and keep the function as is
def archivar_expediente(request):
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        return JsonResponse({'success': False, 'error': 'Sesión no iniciada'})
    
    empleado = Empleado.objects.get(id=empleado_id)
    if not empleado.es_administrador:
        return JsonResponse({'success': False, 'error': 'No tienes permisos para realizar esta acción'})

    if request.method == 'POST':
        try:
            expedientes_ids = request.POST.getlist('expedientes_ids[]')
            motivo = request.POST.get('motivo')
            
            if not expedientes_ids:
                return JsonResponse({'success': False, 'error': 'No se seleccionaron expedientes'})
            
            archivados_count = 0
            for expediente_id in expedientes_ids:
                try:
                    expediente = DosMilSiete.objects.get(id_expediente=expediente_id)
                    
                    # Create archive record with all fields as strings
                    Archivados.objects.create(
                        expediente=expediente.expediente,
                        junta=expediente.junta,
                        actor=expediente.actor,
                        demandado=expediente.demandado,
                        motivo=motivo,
                        fecha_archivo=timezone.now()
                    )
                    
                    # Delete from original table
                    expediente.delete()
                    archivados_count += 1
                except DosMilSiete.DoesNotExist:
                    continue
            
            return JsonResponse({
                'success': True, 
                'message': f'Se archivaron {archivados_count} expedientes correctamente'
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

def obtener_expedientes_ajax(request):
    """Vista para obtener expedientes mediante AJAX para el modal de archivado"""
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        return JsonResponse({'success': False, 'error': 'Sesión no iniciada'})
    
    empleado = Empleado.objects.get(id=empleado_id)
    if not empleado.es_administrador():
        return JsonResponse({'success': False, 'error': 'No tienes permisos para realizar esta acción'})
    
    expedientes = DosMilSiete.objects.all().values('id_expediente', 'expediente', 'junta', 'actor', 'demandado')
    expedientes_list = list(expedientes)
    
    return JsonResponse({'success': True, 'expedientes': expedientes_list})

# Remove @login_required decorator and keep the function as is
def ver_archivados(request):
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        return redirect('iniciar_sesion')
    
    empleado = Empleado.objects.get(id=empleado_id)
    if not empleado.es_administrador():
        messages.error(request, 'No tienes permisos para ver expedientes archivados')
        return redirect('bienvenida')
    
    archivados = Archivados.objects.all().order_by('-fecha_archivo')
    return render(request, 'ver_archivados.html', {'archivados': archivados, 'empleado': empleado})

def exportar_expedientes_excel(request):
    # Verificar sesión
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        messages.error(request, 'Debes iniciar sesión')
        return redirect('iniciar_sesion')
    
    # Obtener todos los expedientes de la tabla 'expedientes'
    expedientes = ConciliacionExpedientes.objects.all()
    
    # Crear el libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Expedientes"
    
    # Estilos para el encabezado
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Definir las columnas (según los campos del modelo ConciliacionExpedientes)
    headers = [
        'EXPEDIENTE', 'JUNTA', 'TOMO', 'ACTOR(NOMBRE)', 'MUJER (NUMERO)', 'HOMBRE(NUMERO)',
        'DEMANDADO (NOMBRE)', 'MUJER (NUMERO)', 'HOMBRE (NUMERO)', 'PERSONA MORAL (NUMERO)',
        'IEBEM', 'SERVICIOS SALUD', 'PODER EJECUTIVO', 'AYUNTAMIENTOS', 'OTROS ORGANISMOS',
        'NO SE HA NOTIFICADO A LAS PARTES', 'EMPL.NO REALIZADO', 'EMPL. EXHORTO',
        'EXHORTOS SIN ENVIAR', 'EXHORTOS CDMX', 'EXHORTOS FORANEOS', 'C.D.E.',
        'TERCERO LLAMADO A JUICIO', 'O.A.P.', 'DESAHOGO PRUEBAS', 'TESTIMONIAL_FALTA_CITAR',
        'PERICIALES_PARTES', 'PERICIALES', 'INFORME_FALTA_HACER', 'INFORME_FALTA_DESAHOGAR',
        'OTRAS_PRUEBAS', 'ALEGATOS', 'PRUEBA PENDIENTE', 'CIERRE', 'LAUDO DICTADO',
        'ABSOLUTORIO', 'CONDENATORIO', 'MONTO DE CONDENA', 'AUTO EJECUCCION', 'TERCERIA',
        'RECURSO REVISION', 'REMATE', 'INACTIVIDAD 1 ANIO', 'INACTIVIDAD 2 ANIOS O +',
        'AMPARO INDIRECTO', 'AMP. DIRECTO CUMPL.', 'SUSTITUCION PATRONAL', 'NO INTERPUESTA',
        'PRESCRIPCION', 'DEPURACION', 'REGULARIZAR', 'REVISO CAPTURO', 'ID_EXPEDIENTE'
    ]
    
    # Escribir encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        ws.column_dimensions[cell.column_letter].width = 18
    
    # Escribir los datos de los expedientes
    for row_num, exp in enumerate(expedientes, 2):
        ws.cell(row=row_num, column=1).value = exp.expediente or ''
        ws.cell(row=row_num, column=2).value = exp.junta or ''
        ws.cell(row=row_num, column=3).value = exp.tomo or ''
        ws.cell(row=row_num, column=4).value = exp.actor_nombre or ''
        ws.cell(row=row_num, column=5).value = exp.mujer_numero or ''
        ws.cell(row=row_num, column=6).value = exp.hombre_numero or ''
        ws.cell(row=row_num, column=7).value = exp.demandado_nombre or ''
        ws.cell(row=row_num, column=8).value = exp.mujer_numero_0 or ''
        ws.cell(row=row_num, column=9).value = exp.hombre_numero_0 or ''
        ws.cell(row=row_num, column=10).value = exp.persona_moral_numero or ''
        ws.cell(row=row_num, column=11).value = exp.iebem or ''
        ws.cell(row=row_num, column=12).value = exp.servicios_salud or ''
        ws.cell(row=row_num, column=13).value = exp.poder_ejecutivo or ''
        ws.cell(row=row_num, column=14).value = exp.ayuntamientos or ''
        ws.cell(row=row_num, column=15).value = exp.otros_organismos or ''
        ws.cell(row=row_num, column=16).value = exp.no_se_ha_notificado_a_las_partes or ''
        ws.cell(row=row_num, column=17).value = exp.empl_no_realizado or ''
        ws.cell(row=row_num, column=18).value = exp.empl_exhorto or ''
        ws.cell(row=row_num, column=19).value = exp.exhortos_sin_enviar or ''
        ws.cell(row=row_num, column=20).value = exp.exhortos_cdmx or ''
        ws.cell(row=row_num, column=21).value = exp.exhortos_foraneos or ''
        ws.cell(row=row_num, column=22).value = exp.cde or ''
        ws.cell(row=row_num, column=23).value = exp.tercero_llamado_a_juicio or ''
        ws.cell(row=row_num, column=24).value = exp.oap or ''
        ws.cell(row=row_num, column=25).value = exp.desahogo_pruebas or ''
        ws.cell(row=row_num, column=26).value = exp.test_falta_citar or ''
        ws.cell(row=row_num, column=27).value = exp.periciales_partes or ''
        ws.cell(row=row_num, column=28).value = exp.pericial_tercero or ''
        ws.cell(row=row_num, column=29).value = exp.inf_falta_hacer or ''
        ws.cell(row=row_num, column=30).value = exp.inf_falta_desahogar or ''
        ws.cell(row=row_num, column=31).value = exp.otras_pruebas or ''
        ws.cell(row=row_num, column=32).value = exp.alegatos or ''
        ws.cell(row=row_num, column=33).value = exp.prueba_pendiente or ''
        ws.cell(row=row_num, column=34).value = exp.cierre or ''
        ws.cell(row=row_num, column=35).value = exp.laudo_dictado or ''
        ws.cell(row=row_num, column=36).value = exp.absolutorio or ''
        ws.cell(row=row_num, column=37).value = exp.condenatorio or ''
        ws.cell(row=row_num, column=38).value = exp.monto or ''
        ws.cell(row=row_num, column=39).value = exp.auto_ejecucion or ''
        ws.cell(row=row_num, column=40).value = exp.terceria or ''
        ws.cell(row=row_num, column=41).value = exp.recurso_revision or ''
        ws.cell(row=row_num, column=42).value = exp.remate or ''
        ws.cell(row=row_num, column=43).value = exp.inactividad_1_anio or ''
        ws.cell(row=row_num, column=44).value = exp.inactividad_2_anios_mas or ''
        ws.cell(row=row_num, column=45).value = exp.amparo_indirecto or ''
        ws.cell(row=row_num, column=46).value = exp.amparo_directo_cumpl or ''
        ws.cell(row=row_num, column=47).value = exp.sustitucion_patronal or ''
        ws.cell(row=row_num, column=48).value = exp.no_interpuesta or ''
        ws.cell(row=row_num, column=49).value = exp.prescripcion or ''
        ws.cell(row=row_num, column=50).value = exp.depuracion or ''
        ws.cell(row=row_num, column=51).value = exp.regularizar or ''
        ws.cell(row=row_num, column=52).value = exp.reviso_capturo or ''
        ws.cell(row=row_num, column=53).value = exp.id_expediente
    
    # Preparar la respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    fecha_actual = datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename=expedientes_{fecha_actual}.xlsx'
    
    # Guardar el libro en la respuesta
    wb.save(response)
    
    return response

